import openpyxl
from decimal import Decimal



def get_spreadsheet(file):
    spreadsheetList = []
    wb = openpyxl.load_workbook(file)
    sheet = wb['Sheet']
    for i in range(1, 20):
        colOne = str(sheet.cell(row=i, column=1).value)
        colTwo = str(sheet.cell(row=i, column=2).value)
        colThree = str(sheet.cell(row=i, column=3).value)
        spreadsheetList.append(str(colOne +  " " + colTwo + " " + colThree ) )

    finalList = '\n'.join(spreadsheetList)
    return finalList

def update_spreadsheet(file, produce, price):

    newPrice = Decimal(price)
    finalPrice = "%.2f" % newPrice

    spreadsheetList = []
    wb = openpyxl.load_workbook(file)
    sheet = wb['Sheet']

    for rowNum in range(2, sheet.max_row):
        produceName = sheet.cell(row=rowNum, column=1).value
        if produceName == produce:
            sheet.cell(row=rowNum, column=2).value = finalPrice

    wb.save(file)

    for i in range(1, 20):
        colOne = str(sheet.cell(row=i, column=1).value)
        colTwo = str(sheet.cell(row=i, column=2).value)
        colThree = str(sheet.cell(row=i, column=3).value)
        spreadsheetList.append(str(colOne +  " " + colTwo + " " + colThree ) )

    finalList = '\n'.join(spreadsheetList)
    return finalList


